'''
relatorio.py

modulo que estará contendo a classe que faz a manipulação do
relatorio criando e colocando as informações sobre faturamento.

'''

#importações para que consiga importar desde a raiz do projeto
import sys
import os

PROJECT_ROOT = os.path.dirname(os.path.dirname(__file__))
sys.path.insert(0,PROJECT_ROOT)

#importações para criação e customização da planilha
import openpyxl
from openpyxl.styles import NamedStyle,Alignment,Font,PatternFill

#importação para tipagem
from model.model_relatorio import ModelPro
from datetime import datetime

#importação para funcionamento da classe
from utils.utils import UtilsPro
from config.logger import LoggerPro

class RelatorioPro:
    """
    Classe responsável pela criação do arquivo de 
    relatório, cria a planilha e acrescenta os dados

    """
    def __init__(self):
        ''''
        Inicializa a classe RelatorioPro.
        
        Cria uma instância da classe UtilsPro
        '''
        self.utils = UtilsPro()
        self.logger = LoggerPro()
        self.workbook = None

    
    def mensal(self, dados_mes: list[ModelPro], dados_mes_anterior: list[ModelPro], total_periodo: dict[str,dict[str,int]]):
        """
        Gera o relatório mensal, contendo o faturamento do 
        mês atual e o faturamento até o mês anterior

        param:
            dados_mes (list[ModelPro]): lista contendo os dados
            do mês escolhido que estão na instância
            dados_mes_anterior (list[ModelPro]): lista contendo os dados
            do mês anterior ao escolhido que estão na instância
            data_atual(datetime): data da geração do relatório
        """

        self.workbook = openpyxl.Workbook()

        # Criando um estilo para dinheiro (R$)
        self.real_style = NamedStyle(name="real_style")
        self.real_style.number_format = '#,##0.00'
        self.workbook.add_named_style(self.real_style)
        
        #ativando a planilha
        self.planilha = self.workbook.active
        self.planilha.title = 'Previsão de Faturamento'
        
        try:
            #pegando mes e ano dos dois relatórios
            mes = self.utils.MESES[dados_mes[0].mes_ref]
            ano = dados_mes[0].ano_ref
            mes_anterior = self.utils.MESES[dados_mes[0].mes_ref - 1]
            comparar_ano_anterior = dados_mes[0].mes_ref == 0
        except IndexError:
            self.logger.mensagem_error("Nenhum dado pra gerar relatório")
            return

        #criando o relatório do mes escolhido
        linha_atual = self.planilha.min_row
        self.planilha.merge_cells(f'A{linha_atual}:F{linha_atual}')
        self.planilha[f'A{linha_atual}'].value = f'Previsão de Faturamento {mes}/{ano}'.upper()
        self.planilha[f'A{linha_atual}'].alignment = Alignment(horizontal="center")
        self.planilha[f'A{linha_atual}'].font = Font(bold=True)
        self.planilha[f'A{linha_atual}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        #adicionando as colunas
        self.planilha.append(self.utils.colunas_all_planilha)
        
        #pegando as posições
        linha_atual = self.planilha.max_row
        coluna_inicio = self.planilha.min_column
        coluna_fim = self.planilha.max_column

        #acrescentando font as colunas
        self.utils.style_font_planilha(self.planilha,linha_atual,linha_atual, coluna_inicio,coluna_fim)
        
        #pegando a posição
        linha_atual = self.planilha.max_row
        
        faturado = 0
        nao_faturado = 0
        
        for dado in dados_mes:
            if dado.data_fat == '-':
                nao_faturado += dado.valor
            else:
                faturado += dado.valor
            
            dado_temp = dado.__dict__.copy()
            dado_temp.pop('tipo')
            dado_temp.pop('mes_ref')
            dado_temp.pop("ano_ref")

            self.planilha.append(list(dado_temp.values()))
        
        #pegando posição
        linha_final = self.planilha.max_row

        #finalizando o relatório
        self.planilha.append(['\n'])
        self.planilha.append(['TOTAL',f"=SUM(B{linha_atual+1}:B{linha_final})"])
        self.planilha.append(['FATURADO', faturado])
        self.planilha.append(['A FATURAR', nao_faturado])
        
        #customizando o relatório
        linha_final = self.planilha.max_row
        self.utils.style_real_planilha(self.planilha,linha_atual+1,linha_final, 2,2, self.real_style)
        self.utils.style_alinhar_planilha(self.planilha,linha_atual+1,linha_final, 3,5)
        self.utils.style_font_planilha(self.planilha,linha_final-2,linha_final, coluna_inicio,2)
        self.utils.style_borda_planilha(self.planilha,1,linha_final, coluna_inicio,coluna_fim)
        self.utils.ajustar_colunas_planilha(self.planilha,linha_atual,linha_final, coluna_inicio,coluna_fim)

        self.planilha.append(['\n'])

        #pegando posição
        linha_atual = self.planilha.max_row + 1

        #criando o relatório do mes anterior
        self.planilha.merge_cells(f'A{linha_atual}:F{linha_atual}')

        #verificando se teve virada de ano
        if comparar_ano_anterior:
            self.planilha[f'A{linha_atual}'].value = f'Faturamento Meses Anteriores {mes_anterior}/{ano - 1}'.upper()
        else:
            self.planilha[f'A{linha_atual}'].value = f'Faturamento Meses Anteriores {mes_anterior}/{ano}'.upper()

        self.planilha[f'A{linha_atual}'].alignment = Alignment(horizontal="center")
        self.planilha[f'A{linha_atual}'].font = Font(bold=True)
        self.planilha[f'A{linha_atual}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        #acrescentando as colunas
        self.planilha.append(self.utils.colunas_all_planilha)
        
        #pegando posição
        linha_atual = self.planilha.max_row

        #colocando font as colunas
        self.utils.style_font_planilha(self.planilha,linha_atual,linha_atual, coluna_inicio,coluna_fim)

        #separando os dados faturados, e pagos
        for dado in dados_mes_anterior:
            dado_temp = dado.__dict__.copy()
            dado_temp.pop('tipo')
            dado_temp.pop('mes_ref')
            dado_temp.pop("ano_ref")

            self.planilha.append(list(dado_temp.values()))
        
        #pegando posição
        linha_final = self.planilha.max_row
        #finalizando o relatório
        self.planilha.append(['\n'])  
        self.planilha.append(['TOTAL',f"=SUM(B{linha_atual+1}:B{linha_final})"])

        linha_final = self.planilha.max_row

        #customizando relatório
        self.utils.style_real_planilha(self.planilha,linha_atual+1,linha_final, 2,2, self.real_style)
        self.utils.style_alinhar_planilha(self.planilha,linha_atual+1,linha_final, 3,5)
        self.utils.style_font_planilha(self.planilha,linha_final,linha_final, coluna_inicio,2)
        self.utils.style_borda_planilha(self.planilha,linha_atual-1,linha_final, coluna_inicio,coluna_fim)

        self.planilha.append(['\n'])

        linha_atual = self.planilha.max_row
        
        dicionario_temp = {'PAGOS 1-10': 0, 'PAGOS 11-20': 0, 'PAGOS 21-31': 0}
        
        for tipo, dicionario in total_periodo.items():
            for periodo,valor in dicionario.items():
                dicionario_temp[periodo] += valor
        
        for periodo, valor in dicionario_temp.items():
            self.planilha.append([f'TOTAL {periodo}',valor])
        
        linha_final = self.planilha.max_row

        #customizando relatório
        self.utils.style_real_planilha(self.planilha,linha_atual+1,linha_final, 2,2, self.real_style)
        self.utils.style_alinhar_planilha(self.planilha,linha_atual+1,linha_final, 3,5)
        self.utils.style_font_planilha(self.planilha,linha_atual+1,linha_final, coluna_inicio,2)
        self.utils.style_borda_planilha(self.planilha,linha_atual+1,linha_final, coluna_inicio,coluna_fim)
        
        self.mensal_separado(dados_mes,dados_mes_anterior,total_periodo)
        
        #nomeando arquivo
        nome = f'Relatorio_{mes}_{ano}.xlsx'

        #salvando a planilha
        self.workbook.save(nome)
    
    def mensal_separado(self, dados_mes: list[ModelPro], dados_mes_anterior: list[ModelPro], total_periodo: dict[str,dict[str,int]]):
        #ativando a planilha
        self.planilha_separada = self.workbook.create_sheet(title='Previsão Separada')

        #pegando mes e ano dos dois relatórios
        mes = self.utils.MESES[dados_mes[0].mes_ref]
        ano = dados_mes[0].ano_ref
        mes_anterior = self.utils.MESES[dados_mes[0].mes_ref - 1]
        comparar_ano_anterior = dados_mes[0].mes_ref == 0

        lista_proprio = []
        lista_consorcio = []

        #separando consocio do próprio
        for dado in dados_mes:
            if dado.tipo == "Próprio":
                lista_proprio.append(dado)
            else:
                lista_consorcio.append(dado)
        
        dicionario = {"Próprio": lista_proprio, "Consórcio": lista_consorcio}

        total_faturado = 0 
        total_a_faturar = 0
        for tipo, dados in dicionario.items():
            
            #criando o relatório do mes escolhido separando consocio do próprio
            linha_atual = self.planilha_separada.max_row
            self.planilha_separada.merge_cells(f'A{linha_atual}:F{linha_atual}')
            self.planilha_separada[f'A{linha_atual}'].value = f'Previsão de Faturamento {tipo} {mes}/{ano}'.upper()
            self.planilha_separada[f'A{linha_atual}'].alignment = Alignment(horizontal="center")
            self.planilha_separada[f'A{linha_atual}'].font = Font(bold=True)
            self.planilha_separada[f'A{linha_atual}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            #adicionando as colunas
            self.planilha_separada.append(self.utils.colunas_all_planilha)
            
            #pegando as posições
            linha_atual = self.planilha_separada.max_row
            coluna_inicio = self.planilha_separada.min_column
            coluna_fim = self.planilha_separada.max_column

            #acrescentando font as colunas
            self.utils.style_font_planilha(self.planilha_separada,linha_atual,linha_atual, coluna_inicio,coluna_fim)
            
            #pegando a posição
            linha_atual = self.planilha_separada.max_row
            
            faturado = 0
            nao_faturado = 0
            
            #inserindo dados na planilha
            for dado in dados:
                #pegando o faturado e não faturado
                if dado.data_fat == '-':
                    nao_faturado += dado.valor
                else:
                    faturado += dado.valor

                dado_temp = dado.__dict__.copy()
                dado_temp.pop('tipo')
                dado_temp.pop('mes_ref')
                dado_temp.pop("ano_ref")

                self.planilha_separada.append(list(dado_temp.values()))
                
                
            
            #pegando posição
            linha_final = self.planilha_separada.max_row

            #finalizando o relatório
            self.planilha_separada.append(['\n'])
            self.planilha_separada.append(['TOTAL',f"=SUM(B{linha_atual+1}:B{linha_final})"])
            self.planilha_separada.append(['FATURADO', faturado])
            self.planilha_separada.append(['A FATURAR', nao_faturado])

            #somando para ver o total consorcio + próprio
            total_faturado += faturado
            total_a_faturar += nao_faturado
            
            #customizando o relatório
            linha_final = self.planilha_separada.max_row
            self.utils.style_real_planilha(self.planilha_separada,linha_atual+1,linha_final, 2,2, self.real_style)
            self.utils.style_alinhar_planilha(self.planilha_separada,linha_atual+1,linha_final, 3,5)
            self.utils.style_font_planilha(self.planilha_separada,linha_final-2,linha_final, coluna_inicio,2)
            self.utils.style_borda_planilha(self.planilha_separada,linha_atual-1,linha_final, coluna_inicio,coluna_fim)
            self.utils.ajustar_colunas_planilha(self.planilha_separada,linha_atual,linha_final, coluna_inicio,coluna_fim)

            self.planilha_separada.append(['\n'])
            self.planilha_separada.append(['\n'])
        
        #pegando posição
        linha_atual = self.planilha_separada.max_row

        self.planilha_separada.append(['TOTAL FATURADO', total_faturado])
        self.planilha_separada.append(['TOTAL A FATURAR', total_a_faturar])

        #pegando posição
        linha_final = self.planilha_separada.max_row

        #customizando parte final do relatório
        self.utils.style_real_planilha(self.planilha_separada,linha_atual+1,linha_final, 2,2, self.real_style)
        self.utils.style_font_planilha(self.planilha_separada,linha_final-2,linha_final, coluna_inicio,2)
        self.utils.style_borda_planilha(self.planilha_separada,linha_atual+1,linha_final, coluna_inicio,coluna_fim)

        self.planilha_separada.append(['\n'])


        lista_proprio = []
        lista_consorcio = []

        #separando consorcio do próprio
        for dado in dados_mes_anterior:
            if dado.tipo == "Próprio":
                lista_proprio.append(dado)
            else:
                lista_consorcio.append(dado)
        
        dicionario = {"Próprio": lista_proprio, "Consórcio": lista_consorcio}

        for tipo, dados in dicionario.items():
            #pegando posição
            linha_atual = self.planilha_separada.max_row + 1

            #criando o relatório do mes anterior
            self.planilha_separada.merge_cells(f'A{linha_atual}:F{linha_atual}')

            #verificando se teve virada de ano
            if comparar_ano_anterior:
                self.planilha_separada[f'A{linha_atual}'].value = f'Faturamento Meses Anteriores {tipo} {mes_anterior}/{ano - 1}'.upper()
            else:
                self.planilha_separada[f'A{linha_atual}'].value = f'Faturamento Meses Anteriores {tipo} {mes_anterior}/{ano}'.upper()

            self.planilha_separada[f'A{linha_atual}'].alignment = Alignment(horizontal="center")
            self.planilha_separada[f'A{linha_atual}'].font = Font(bold=True)
            self.planilha_separada[f'A{linha_atual}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            #acrescentando as colunas
            self.planilha_separada.append(self.utils.colunas_all_planilha)
            
            #pegando posição
            linha_atual = self.planilha_separada.max_row

            #colocando font as colunas
            self.utils.style_font_planilha(self.planilha_separada,linha_atual,linha_atual, coluna_inicio,coluna_fim)

            #inserindo dados na planilha
            for dado in dados:
                
                dado_temp = dado.__dict__.copy()
                dado_temp.pop('tipo')
                dado_temp.pop('mes_ref')
                dado_temp.pop("ano_ref")

                self.planilha_separada.append(list(dado_temp.values()))
            
            #pegando posição
            linha_final = self.planilha_separada.max_row
            
            #finalizando o relatório
            self.planilha_separada.append(['\n'])  
            self.planilha_separada.append(['TOTAL',f"=SUM(B{linha_atual+1}:B{linha_final})"])
            
            #pegando posição
            linha_final = self.planilha_separada.max_row

            #customizando relatório
            self.utils.style_real_planilha(self.planilha_separada,linha_atual+1,linha_final, 2,2, self.real_style)
            self.utils.style_alinhar_planilha(self.planilha_separada,linha_atual+1,linha_final, 3,5)
            self.utils.style_font_planilha(self.planilha_separada,linha_final,linha_final, coluna_inicio,2)
            self.utils.style_borda_planilha(self.planilha_separada,linha_atual-1,linha_final, coluna_inicio,coluna_fim)
            self.utils.ajustar_colunas_planilha(self.planilha_separada,linha_atual,linha_final, coluna_inicio,coluna_fim)

            self.planilha_separada.append(['\n'])
            
        
            linha_atual = self.planilha_separada.max_row

            for periodo, valor in total_periodo[tipo].items():
                self.planilha_separada.append([f'TOTAL {periodo}',valor])
            
            linha_final = self.planilha_separada.max_row

            #customizando relatório
            self.utils.style_real_planilha(self.planilha_separada,linha_atual+1,linha_final, 2,2, self.real_style)
            self.utils.style_alinhar_planilha(self.planilha_separada,linha_atual+1,linha_final, 3,5)
            self.utils.style_font_planilha(self.planilha_separada,linha_atual+1,linha_final, coluna_inicio,2)
            self.utils.style_borda_planilha(self.planilha_separada,linha_atual+1,linha_final, coluna_inicio,coluna_fim)
            self.planilha_separada.append(['\n'])

                    









        


        