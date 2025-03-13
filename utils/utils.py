'''
utils.py

modulo que estará contendo a classe com funções auxiliares,
listas, entre outras ajudas para o projeto.

'''
import os
import sys

PROJECT_ROOT = os.path.dirname(os.path.dirname(__file__))
sys.path.insert(0,PROJECT_ROOT)


from datetime import datetime
import customtkinter as ctk

from model.model_nota import ModelNota
from model.model_cliente import ModelCliente
from model.model import ModelPro
from config.logger import LoggerPro
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, NamedStyle, Border, Side,Alignment
from openpyxl.utils import get_column_letter

class UtilsPro:
    """
    Classe utilitária para manipulação e formatação de dados em um sistema de gerenciamento financeiro.

    Contém métodos para formatação de valores monetários, manipulação de datas, limpeza de formulários,
    controle de tela cheia, entre outras funções auxiliares.
    """

    def __init__(self) -> None:
        """
        Inicializa a classe com listas de meses, colunas de dados e tipos de clientes.
        """
        self.MESES = [
            'JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO', 
            'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO'
        ]

        self.colunas_clientes = ["id",'Nome', "Centro de Custo", "Tipo", "Descrição"]

        self.colunas_notas = [
            "id", "Centro de Custo", "Numero da Nota", "Valor da Nota", "Data de Faturamento", 
            "Data de Pagamento", "Mês de Referência", "Ano de Referência"
        ]

        self.colunas_all = [
            'Nome', "Valor da Nota", "Centro de Custo", 
            "Data de Faturamento", "Data de Pagamento", "Descrição"
        ]

        self.colunas_all_planilha = [
            'CLIENTE', "VALOR DA NOTA", "CENTRO DE CUSTO", 
            "FAT", "PAG", "DESCRIÇÃO"
        ]

        self.tipo_cliente = ["Consórcio", "Próprio"]

        self.logger = LoggerPro()

    def formatar_cliente(self, cliente: ModelCliente) -> ModelCliente:
        cliente.nome = str(cliente.nome).strip()
        cliente.cc = str(cliente.cc).strip()
        cliente.tipo = str(cliente.tipo).strip()
        cliente.descricao = str(cliente.descricao).strip()

        return cliente

    def formatar_nota(self, nota: ModelNota) -> ModelNota:
        """
        Formata os valores de um dicionário, convertendo datas e valores monetários.

        param:
            nota(Nota): Dicionário contendo os dados a serem formatados.
        
        return
            (Nota): Dicionário com os dados formatados.
        """
        nota.cc = str(nota.cc).strip()
        nota.numero_nota = str(nota.numero_nota).strip()
        nota.valor_nota = self.formatar_dinheiro(str(nota.valor_nota))
        nota.data_fat = self.formatar_data(str(nota.data_fat))
        nota.data_pag = self.formatar_data(str(nota.data_pag))
        if nota.data_fat != '':
            data = datetime.strptime(nota.data_fat,"%Y-%m-%d")
            nota.mes_ref = data.month - 1
            nota.ano_ref = data.year
        else:    
            nota.mes_ref = self.MESES.index(str(nota.mes_ref).strip())
            nota.ano_ref = int(str(nota.ano_ref).strip())

        return nota
    
    def customizar_nota(self, notas: list[ModelNota]) -> list[ModelNota]:
        """
        Formata os valores de um dicionário, convertendo datas e valores monetários.

        param:
            nota(Nota): Dicionário contendo os dados a serem formatados.
        
        return
            (Nota): Dicionário com os dados formatados.
        """
        lista = []
        for nota in notas:
            nota.valor_nota = self.customizar_dinheiro(str(nota.valor_nota))
            nota.data_fat = self.customizar_data(str(nota.data_fat))
            nota.data_pag = self.customizar_data(str(nota.data_pag))
            nota.mes_ref = self.MESES[int(nota.mes_ref)]
            nota.ano_ref = str(nota.ano_ref)
            lista.append(nota)

        return lista
    
    def customizar_modelo(self, modelo: list[ModelPro]) -> list[ModelPro]:
        """
        Formata os valores de um dicionário, convertendo datas e valores monetários.

        param:
            nota(Nota): Dicionário contendo os dados a serem formatados.
        
        return
            (Nota): Dicionário com os dados formatados.
        """
        lista = []
        for molde in modelo:
            molde.valor = self.customizar_dinheiro(str(molde.valor))
            molde.data_fat = self.customizar_data(str(molde.data_fat))
            molde.data_pag = self.customizar_data(str(molde.data_pag))
            molde.mes_ref = self.MESES[int(molde.mes_ref)]
            molde.ano_ref = str(molde.ano_ref)
            lista.append(molde)

        return lista

    def formatar_dinheiro(self, dinheiro: str) -> float:
        """
        Converte uma string representando dinheiro em um valor float.

        param:
            dinheiro(str): String representando um valor monetário (ex: "R$1.234,56").
        
        return: 
            (float): Valor convertido para float.
        
        raises 
            ValueError: Se a conversão falhar.
        """
        valor = dinheiro.strip()
        if valor == "":
            self.logger.mensagem_error("Valor da Nota não pode estar em branco")
            raise ValueError("Valor da Nota não pode estar em branco")

        try:
            valor = valor.replace('R$', '').replace('.', '').replace(',', '.')
            return float(valor.strip())
        except ValueError:
            self.logger.mensagem_error("Erro ao converter valor da nota verifique se existe uma letra presente")
            raise ValueError("Erro ao converter valor da nota verifique se existe uma letra presente")
    
    def customizar_dinheiro(self, dinheiro: str) -> float:
        """
        Converte uma string representando dinheiro em um valor float.

        param:
            dinheiro(str): String representando um valor monetário (ex: "R$1.234,56").
        
        return: 
            (float): Valor convertido para float.
        
        raises 
            ValueError: Se a conversão falhar.
        """
        valor = dinheiro.strip()
        valor = valor.replace('.',',')
        valor = valor.split(',')
        valor[0] = valor[0][::-1]
        valor_customizado = ''
        contador = 0

        for number in valor[0]:
            if contador // 3 == 1:
                valor_customizado +=  '.' + number
                contador = 1
                continue
            
            valor_customizado += number
            contador += 1
        
        valor_customizado = valor_customizado[::-1] 
        if len(valor) != 1:
            valor_customizado += ","+ valor[1]

        return valor_customizado

    def formatar_data(self, data: str) -> str:
        """
        Converte uma data no formato brasileiro para o formato internacional (YYYY-MM-DD) e vice-versa.

        param:
            data(str): String representando uma data (ex: "10/05/2023" ou "2023-05-10").
        
        return: 
            (str): Data formatada como string.
        """
        try:
            data = data.strip()
            if data == '':
                return data
            else:
                data_for = datetime.strptime(data,"%d/%m/%Y")
                data_for = data_for.strftime("%Y-%m-%d")

                return data_for
        except ValueError:
            try:
                data_for = datetime.strptime(data,"%Y-%m-%d")
                data_for = data_for.strftime("%Y-%m-%d")
                
                return data_for
            except ValueError:
                self.logger.mensagem_error("Formato da Data invalido, use somente DD/MM/AA ou AA-MM-DD")
                raise ValueError
    
    def customizar_data(self, data: str) -> str:
        """
        Converte uma data no formato brasileiro para o formato internacional (YYYY-MM-DD) e vice-versa.

        param:
            data(str): String representando uma data (ex: "10/05/2023" ou "2023-05-10").
        
        return: 
            (str): Data formatada como string.
        """
        if data == '':
            return data
        try:
            data_for = datetime.strptime(data,"%Y-%m-%d")
            data_for = data_for.strftime("%d/%m/%Y")
        except ValueError:
            data_for = datetime.strptime(data,"%d/%m/%Y")
            data_for = data_for.strftime("%d/%m/%Y")
        
        return data_for
        

    
    def validar_data(self,nota: ModelNota) -> bool:
        if nota.data_pag == '':
            return True
        try:
            if datetime.strptime(nota.data_fat,"%Y-%m-%d") > datetime.strptime(nota.data_pag,"%Y-%m-%d"):
                return False
        except ValueError:
            try:
                if datetime.strptime(nota.data_fat,"%d/%m/%Y") > datetime.strptime(nota.data_pag,"%d/%m/%Y"):
                    return False
            except ValueError:
                self.logger.mensagem_error("Formato da Data invalido, use somente DD/MM/AA ou AA-MM-DD")
                return False
        
        return True
        

    def apagar_valores(self, dicionario: dict, quant: int, cliente: bool = False, nota: bool = False) -> None:
        """
        Limpa os valores dos campos de entrada conforme a quantidade de registros.

        param:
            dicionario(dict): Dicionário contendo os campos de entrada.
            quant(int): Quantidade de registros.
            cliente(bool): Se True, limpa campos relacionados a clientes.
            nota(bool): Se True, limpa campos relacionados a notas fiscais.
        """
        for chave, dado in dicionario.items():
            match chave:
                case 'Ano de Referência':
                    if quant == 1:
                        dado.delete(0, 'end')
                        dado.insert(0, str(self.pegar_ano_atual()))
                case "Mês de Referência":
                    if quant == 1:
                        dado.set(self.MESES[self.pegar_mes_atual()])
                case "Quantidade de Registros":
                    if quant > 1:
                        dado.set(str(quant - 1))
                case "Tipo de Cliente":
                    continue
                case 'Centro de Custo':
                    if quant > 1 and nota:
                        continue
                    elif cliente:
                        dado.delete(0, 'end')
                    else:
                        dado.delete(0, 'end')
                case _:
                    if quant == 1 or nota:
                        dado.delete(0, 'end')

    def limpar(self, janela: ctk.CTk) -> None:
        """
        Remove todos os widgets da janela, exceto os primeiros botões principais.

        param 
            janela(CTk): A janela de onde os widgets serão removidos.
        """
        contador = 0
        for componente in janela.winfo_children():
            if contador > 2 and isinstance(componente, (ctk.CTkLabel, ctk.CTkEntry,ctk.CTkButton, ctk.CTkComboBox)):
                componente.destroy()
                
            contador += 1

    def pegar_mes_atual(self) -> int:
        """
        Retorna o índice do mês atual (baseado em zero, onde janeiro é 0 e dezembro é 11).

        return: 
            (int): Índice do mês atual.
        """
        return datetime.now().month - 1

    def pegar_ano_atual(self) -> int:
        """
        Retorna o ano atual.

        return: 
            (int): Ano atual como inteiro.
        """
        return datetime.now().year
    
    def style_font_planilha(self,planilha: Worksheet, inicio_linha: int, fim_linha: int, inicio_coluna: int, fim_coluna: int):
        for row in planilha.iter_rows(min_row=inicio_linha, max_row=fim_linha, min_col=inicio_coluna, max_col=fim_coluna):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(bold=True)
    
    def style_alinhar_planilha(self,planilha: Worksheet, inicio_linha: int, fim_linha: int, inicio_coluna: int, fim_coluna: int):
        for row in planilha.iter_rows(min_row=inicio_linha, max_row=fim_linha, min_col=inicio_coluna, max_col=fim_coluna):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

    def style_real_planilha(self,planilha: Worksheet, inicio_linha: int, fim_linha: int, inicio_coluna: int, fim_coluna: int, style_real: NamedStyle):
        for row in planilha.iter_rows(min_row=inicio_linha, max_row=fim_linha, min_col=inicio_coluna, max_col=fim_coluna):
            for cell in row:
                cell.style =style_real
    
    def ajustar_colunas_planilha(self,planilha: Worksheet):
        for col in planilha.iter_cols():
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            planilha.column_dimensions[col_letter].width = max_length + 2
    
    def style_borda_planilha(self,planilha: Worksheet, inicio_linha: int, fim_linha: int, inicio_coluna: int, fim_coluna: int) -> None:

        # Definir um estilo de borda em negrito
        borda_negrito = Border(
            left=Side(style="thin"),   # Borda esquerda grossa
            right=Side(style="thin"),  # Borda direita grossa
            top=Side(style="thin"),    # Borda superior grossa
            bottom=Side(style="thin")  # Borda inferior grossa
        )
        
        for row in planilha.iter_rows(min_row=inicio_linha, max_row=fim_linha, min_col=inicio_coluna, max_col=fim_coluna):
            for cell in row:
                cell.border = borda_negrito