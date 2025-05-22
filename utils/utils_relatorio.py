"""
utils.py

modulo que estará contendo a classe com funções auxiliares,
listas, entre outras ajudas para o projeto.

"""

import os
import sys
from datetime import datetime

PROJECT_ROOT = os.path.dirname(os.path.dirname(__file__))
sys.path.insert(0, PROJECT_ROOT)


from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, NamedStyle, Border, Side, Alignment

from app.model.model_relatorio import ModelRelatorio


class UtilsRelatorio:
    def __init__(self):
        self.colunas_all = [
            "Nome",
            "Valor da Nota",
            "Centro de Custo",
            "Data de Faturamento",
            "Data de Pagamento",
            "Descrição",
        ]

    def customizar_modelo(self, modelo: list[ModelRelatorio]) -> list[ModelRelatorio]:
        """
        Formata os valores de um dicionário, convertendo datas e valores monetários.

        param:
            nota(Nota): Dicionário contendo os dados a serem formatados.

        return
            (Nota): Dicionário com os dados formatados.
        """
        lista = []
        for molde in modelo:
            molde.valor = self.customizar_dinheiro(str(molde.valor))
            molde.data_fat = self.customizar_data(str(molde.data_fat))
            molde.data_pag = self.customizar_data(str(molde.data_pag))
            molde.mes_ref = self.MESES[int(molde.mes_ref)]
            molde.ano_ref = str(molde.ano_ref)
            lista.append(molde)

        return lista

    def customizar_dinheiro(self, dinheiro: str) -> float:
        """
        Converte uma string representando dinheiro em um valor float.

        param:
            dinheiro(str): String representando um valor monetário (ex: "R$1.234,56").

        return:
            (float): Valor convertido para float.

        raises
            ValueError: Se a conversão falhar.
        """
        valor = dinheiro.strip()
        valor = valor.replace(".", ",")
        valor = valor.split(",")
        valor[0] = valor[0][::-1]
        valor_customizado = ""
        contador = 0

        for number in valor[0]:
            if contador // 3 == 1:
                valor_customizado += "." + number
                contador = 1
                continue

            valor_customizado += number
            contador += 1

        valor_customizado = valor_customizado[::-1]
        if len(valor) != 1:
            valor_customizado += "," + valor[1]

        return valor_customizado

    def customizar_data(self, data: str) -> str:
        """
        Converte uma data no formato brasileiro para o formato internacional (YYYY-MM-DD) e vice-versa.

        param:
            data(str): String representando uma data (ex: "10/05/2023" ou "2023-05-10").

        return:
            (str): Data formatada como string.
        """
        if data == "":
            return data
        try:
            data_for = datetime.strptime(data, "%Y-%m-%d")
            data_for = data_for.strftime("%d/%m/%Y")
        except ValueError:
            data_for = datetime.strptime(data, "%d/%m/%Y")
            data_for = data_for.strftime("%d/%m/%Y")

        return data_for

    def style_font_planilha(
        self,
        planilha: Worksheet,
        inicio_linha: int,
        fim_linha: int,
        inicio_coluna: int,
        fim_coluna: int,
    ):
        for row in planilha.iter_rows(
            min_row=inicio_linha,
            max_row=fim_linha,
            min_col=inicio_coluna,
            max_col=fim_coluna,
        ):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.font = Font(bold=True)

    def style_alinhar_planilha(
        self,
        planilha: Worksheet,
        inicio_linha: int,
        fim_linha: int,
        inicio_coluna: int,
        fim_coluna: int,
    ):
        for row in planilha.iter_rows(
            min_row=inicio_linha,
            max_row=fim_linha,
            min_col=inicio_coluna,
            max_col=fim_coluna,
        ):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

    def style_real_planilha(
        self,
        planilha: Worksheet,
        inicio_linha: int,
        fim_linha: int,
        inicio_coluna: int,
        fim_coluna: int,
        style_real: NamedStyle,
    ):
        for row in planilha.iter_rows(
            min_row=inicio_linha,
            max_row=fim_linha,
            min_col=inicio_coluna,
            max_col=fim_coluna,
        ):
            for cell in row:
                cell.style = style_real

    def ajustar_colunas_planilha(
        self,
        planilha: Worksheet,
        inicio_linha: int,
        fim_linha: int,
        inicio_coluna: int,
        fim_coluna: int,
    ):
        for col in planilha.iter_cols(
            inicio_coluna, fim_coluna, inicio_linha, fim_linha
        ):
            col_letter = col[0].column_letter
            if col_letter in ["D", "E"]:
                max_length = 0
            else:
                max_length = planilha.column_dimensions[col_letter].width
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            planilha.column_dimensions[col_letter].width = max_length + 2

    def style_borda_planilha(
        self,
        planilha: Worksheet,
        inicio_linha: int,
        fim_linha: int,
        inicio_coluna: int,
        fim_coluna: int,
    ) -> None:

        # Definir um estilo de borda em negrito
        borda_negrito = Border(
            left=Side(style="thin"),  # Borda esquerda grossa
            right=Side(style="thin"),  # Borda direita grossa
            top=Side(style="thin"),  # Borda superior grossa
            bottom=Side(style="thin"),  # Borda inferior grossa
        )

        for row in planilha.iter_rows(
            min_row=inicio_linha,
            max_row=fim_linha,
            min_col=inicio_coluna,
            max_col=fim_coluna,
        ):
            for cell in row:
                cell.border = borda_negrito
